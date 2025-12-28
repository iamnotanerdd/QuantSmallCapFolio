import pandas as pd
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Configuration
# INPUT_FOLDER = os.path.join(os.path.dirname(__file__), "2025_Disclosures") # OLD
INPUT_FOLDER = os.path.join(os.path.dirname(__file__), "Disclosures")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "Quant_SmallCap_Equity_Analysis.xlsx")
SCHEME_NAME_TITLE = "Quant Small Cap Fund"

# Column Mapping based on Quant Excel Structure
# Header found around Row 7 usually
# SR | ISIN | NAME OF THE INSTRUMENT | RATING | INDUSTRY | QUANTITY | MARKET VALUE(Rs.in Lakhs) | % to NAV
COLUMN_MAPPING = {
    "NAME OF THE INSTRUMENT": "Name",
    "ISIN": "ISIN",
    "INDUSTRY": "Rating", # User requested INDUSTRY data in this column
    "QUANTITY": "Quantity",
    "MARKET VALUE(Rs.in Lakhs)": "MarketValue",
    "MARKET VALUE": "MarketValue", # Variation
    "% to NAV": "PctAssets",
    "% TO NAV": "PctAssets"
}

def normalize_header(header):
    if not isinstance(header, str): return ""
    h = header.upper().strip().replace('\n', ' ')
    
    # Direct matching or keyword matching
    if "NAME OF THE INSTRUMENT" in h: return "Name"
    if "ISIN" in h: return "ISIN"
    if "INDUSTRY" in h: return "Rating" # Map INDUSTRY to Rating key
    if "QUANTITY" in h: return "Quantity"
    if "MARKET VALUE" in h: return "MarketValue"
    if "% TO NAV" in h or "% TO NET ASSETS" in h: return "PctAssets"
    
    return h

def read_portfolio_file(filepath):
    try:
        # 1. Find Header Row
        # Read first 15 rows to find "ISIN"
        df_scan = pd.read_excel(filepath, nrows=15, header=None)
        header_idx = -1
        
        for idx, row in df_scan.iterrows():
            row_vals = [str(x).upper() for x in row.values]
            if "ISIN" in row_vals and ("QUANTITY" in row_vals or "QTY" in row_vals):
                header_idx = idx
                break
        
        if header_idx == -1:
            print(f"Could not find header in {os.path.basename(filepath)}")
            return None
            
        # 2. Read Data
        df = pd.read_excel(filepath, header=header_idx)
        
        # 3. Rename Columns
        new_cols = {}
        for col in df.columns:
            norm = normalize_header(col)
            if norm in COLUMN_MAPPING.values():
                new_cols[col] = norm
        
        df.rename(columns=new_cols, inplace=True)
        
        # 4. Filter Rows
        # Quant format usually has "EQUITY & EQUITY RELATED" then data, then "Total"
        # We need rows where ISIN is present.
        
        filtered_rows = []
        
        # Check if required columns exist
        req_cols = ["Name", "ISIN", "Quantity", "MarketValue"]
        if not all(col in df.columns for col in req_cols):
             print(f"Missing columns in {os.path.basename(filepath)}: Found {df.columns.tolist()}")
             return None

        # Logic to extract Equity section
        # We can just look for valid ISINs that start with 'INE' or 'IN' usually for Indian stocks, 
        # but simpler is just non-null ISIN and non-null Name.
        
        for idx, row in df.iterrows():
            # Stop conditions (Totals etc)
            name_val = str(row.get('Name', ''))
            if "Total" in name_val and "Grand" in name_val:
                break
                
            isin = row.get('ISIN')
            if pd.notna(isin) and str(isin).strip().upper() != 'NAN' and str(isin).strip() != '':
                filtered_rows.append(row)
                
        return pd.DataFrame(filtered_rows)

    except Exception as e:
        print(f"Error processing {os.path.basename(filepath)}: {e}")
        return None

def main():
    print(f"Checking for files in {INPUT_FOLDER}...")
    files = glob.glob(os.path.join(INPUT_FOLDER, "*.xlsx"))
    files += glob.glob(os.path.join(INPUT_FOLDER, "*.xls"))
    
    if not files:
        print("No Excel files found.")
        return
        
    print(f"Found {len(files)} files. Aggregating...")
    
    portfolio = {}
    all_months = set()
    
    # Process files
    for f in files:
        fname = os.path.basename(f)
        # Extract Month_Year from filename: quant_Small_Cap_Fund_Jan_2025.xlsx
        # We want 'Jan_2025' or 'January 2025'. 
        # Logic: Split by '_' and take last 2 parts minus extension
        
        try:
            name_parts = fname.replace(".xlsx", "").replace(".xls", "").split('_')
            # Assuming format: prefix_Month_Year
            # month = name_parts[-2]
            # year = name_parts[-1]
            # label = f"{month} {year}"
            
            # Or just use the filename as key for now, or map 'Jan' -> 'January'
            month_abbr = name_parts[-2]
            year = name_parts[-1]
            
            month_map = {
                'Jan': 'January', 'Feb': 'February', 'Mar': 'March', 'Apr': 'April',
                'May': 'May', 'Jun': 'June', 'Jul': 'July', 'Aug': 'August',
                'Sep': 'September', 'Oct': 'October', 'Nov': 'November', 'Dec': 'December',
                'Sept': 'September'
            }
            
            full_month = month_map.get(month_abbr, month_abbr)
            month_label = f"{full_month} {year}" # e.g. January 2025
            
        except:
             month_label = fname
        
        all_months.add(month_label)
        print(f"Reading {month_label}...")
        
        df = read_portfolio_file(f)
        if df is None: continue
        
        for _, row in df.iterrows():
            isin = str(row['ISIN']).strip()
            
            if isin not in portfolio:
                portfolio[isin] = {
                    'Name': row.get('Name'),
                    'Rating': row.get('Rating'), # This is INDUSTRY now
                    'Months': {}
                }
            
            # Update static data if missing
            if not portfolio[isin]['Name'] and pd.notna(row.get('Name')):
                portfolio[isin]['Name'] = row.get('Name')
            if not portfolio[isin]['Rating'] and pd.notna(row.get('Rating')):
                portfolio[isin]['Rating'] = row.get('Rating')
                
            portfolio[isin]['Months'][month_label] = {
                'Quantity': row.get('Quantity'),
                'MarketValue': row.get('MarketValue'),
                'PctAssets': row.get('PctAssets')
            }

    # Sort Months
    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 
                   'July', 'August', 'September', 'October', 'November', 'December']
                   
    def sort_key(m_str):
        try:
            parts = m_str.split(' ')
            m = parts[0]
            y = int(parts[1])
            return (y, month_order.index(m) if m in month_order else 99)
        except:
            return (9999, 99)
            
    sorted_months = sorted(list(all_months), key=sort_key)
    print(f"Months ordered: {sorted_months}")
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Equity Analysis"
    
    # --- STYLES ---
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    middle_align = Alignment(vertical='center')
    
    thin_border = Side(style='thin')
    med_border = Side(style='medium')
    
    def get_border(col_idx):
        right = thin_border
        # Month blocks of 3. 
        # Static: 1,2,3. End at 3.
        # M1: 4,5,6. End at 6.
        if col_idx >= 3 and (col_idx - 3) % 3 == 0:
            right = med_border
        return Border(left=thin_border, right=right, top=thin_border, bottom=thin_border)

    # 1. Title
    ws['A1'] = SCHEME_NAME_TITLE
    # Merge later
    
    # 2. Static Headers
    static_headers = ["Name of the Instrument", "ISIN", "Industry/Rating"]
    for i, h in enumerate(static_headers, 1):
        c = ws.cell(2, i, h)
        c.font = header_font
        c.alignment = center_align
        c.border = get_border(i)

    # 3. Dynamic Headers
    col_curr = 4
    col_totals = {} # index -> sum
    
    for m in sorted_months:
        # Month Header
        c = ws.cell(2, col_curr, m)
        ws.merge_cells(start_row=2, start_column=col_curr, end_row=2, end_column=col_curr+2)
        c.font = header_font
        c.alignment = center_align
        c.border = get_border(col_curr+2) # Border on the rightmost merged cell? 
        # Merged cells border handling is tricky in openpyxl, need to style the range or just the main cell
        # Usually checking the edges is enough
        
        # Sub Headers
        ws.cell(3, col_curr, "Quantity").font = header_font
        ws.cell(3, col_curr+1, "Market Value (Rs. Lakhs)").font = header_font
        ws.cell(3, col_curr+2, "% Net Assets").font = header_font
        
        for k in range(3):
            cell = ws.cell(3, col_curr+k)
            cell.alignment = center_align
            cell.border = get_border(col_curr+k)
            col_totals[col_curr+k] = 0.0

        col_curr += 3
        
    final_col = col_curr - 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_col)
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # 4. Data
    row_curr = 4
    for isin, data in portfolio.items():
        ws.cell(row_curr, 1, data['Name']).alignment = middle_align
        ws.cell(row_curr, 2, isin).alignment = middle_align
        ws.cell(row_curr, 3, data['Rating']).alignment = middle_align # Industry
        
        # Borders for static
        for i in range(1, 4):
            ws.cell(row_curr, i).border = get_border(i)
        
        c_ptr = 4
        for m in sorted_months:
            m_data = data['Months'].get(m, {})
            
            vals = [
                m_data.get('Quantity', 0),
                m_data.get('MarketValue', 0),
                m_data.get('PctAssets', 0)
            ]
            
            for i, val in enumerate(vals):
                real_col = c_ptr + i
                try: v_float = float(val) if val is not None else 0.0
                except: v_float = 0.0
                
                # Logic to fix Percentage: if column is PctAssets (index 2), divide by 100
                if i == 2:
                    v_float = v_float / 100.0
                
                col_totals[real_col] += v_float
                
                c = ws.cell(row_curr, real_col, v_float if val is not None else None)
                c.alignment = middle_align
                c.border = get_border(real_col)
                
                # Format
                if i == 0: c.number_format = '#,##0'
                elif i == 1: c.number_format = '#,##0.00'
                elif i == 2: c.number_format = '0.00%'

            c_ptr += 3
        row_curr += 1
        
    # 5. Totals
    ws.cell(row_curr, 1, "Total").font = Font(bold=True)
    for c_idx, total in col_totals.items():
        c = ws.cell(row_curr, c_idx, total)
        c.font = Font(bold=True)
        c.border = get_border(c_idx)
        
        rem = (c_idx - 4) % 3
        if rem == 0: c.number_format = '#,##0'
        elif rem == 1: c.number_format = '#,##0.00'
        elif rem == 2: c.number_format = '0.00%'

    # Auto Width
    from openpyxl.utils import get_column_letter
    for col_i in range(1, final_col + 1):
        col_let = get_column_letter(col_i)
        ws.column_dimensions[col_let].width = 15 if col_i > 3 else 30

    wb.save(OUTPUT_FILE)
    print(f"Successfully saved aggregated data to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
